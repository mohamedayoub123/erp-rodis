import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, date, time, timedelta
from pathlib import Path

from openpyxl import load_workbook

SHEET_NAME = "SUIVI TIRAGE"

# Index (0-based) de chaque colonne dans la feuille SUIVI TIRAGE - correspond
# a l'ordre reel des colonnes A..AR de GFPC-ENR-007_SUIVI_PRODUCTION.xlsx.
COL_DATE = 0
COL_ARTICLE = 3
COL_POIDS_MOYEN = 4
COL_CADENCE_NET = 7
COL_ZONE = 8
COL_LIGNE = 9
COL_CHEF_ZONE = 10
COL_CHEF_LIGNE = 11
COL_TIREUR = 12
COL_PROGRAMME_DE_PRODUCTION = 14
COL_CODE_LABO = 15
COL_NBRE_CARTONS = 16
COL_TEMPS_OUVERT = 19
COL_FIN = 21
COL_ARRET_DEPOT = 23
COL_ARRET_CONSOMMABLE = 24
COL_ARRET_MANQUE_CONDITIONNEMENT = 25
COL_ARRET_MANQUE_VRAC = 26
COL_ARRET_TECHNIQUE = 27
COL_ARRET_COUPURE_COURANT = 28
COL_ARRET_RACLAGE_VRAC = 29
COL_ARRET_CHANGEMENT_LOT = 30
COL_ARRET_FLACONS_NC = 31
COL_ARRET_AUTRE = 32
COL_PRODUCTION = 35


def load_env(path: Path) -> None:
    if not path.exists():
        return
    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def request_json(url: str, api_key: str, method: str = "GET", payload=None, extra_headers=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    if extra_headers:
        headers.update(extra_headers)

    request = urllib.request.Request(url, data=body, headers=headers, method=method)

    try:
        with urllib.request.urlopen(request) as response:
            raw = response.read().decode("utf-8", errors="ignore")
            return json.loads(raw) if raw else []
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="ignore")
        raise RuntimeError(f"{error.code} {url}: {detail}") from error


def fetch_all(base_url: str, api_key: str, table: str, select: str) -> list[dict]:
    rows: list[dict] = []
    from_idx = 0
    page_size = 1000

    while True:
        query = urllib.parse.urlencode({"select": select})
        request = urllib.request.Request(
            f"{base_url}/rest/v1/{table}?{query}",
            headers={
                "apikey": api_key,
                "Authorization": f"Bearer {api_key}",
                "Range": f"{from_idx}-{from_idx + page_size - 1}",
            },
        )
        with urllib.request.urlopen(request) as response:
            chunk = json.loads(response.read().decode("utf-8"))

        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        from_idx += page_size

    return rows


def normalize_code(value) -> str:
    if value in (None, ""):
        return ""
    return str(value).replace("\xa0", " ").strip().upper()


# Meme rapprochement tolerant que import_production_data.py : ignore
# tirets/espaces multiples et corrige la faute de frappe connue
# "PRERFECT" (notre base) / "PERFECT" (Excel), pour retrouver l'article par
# son nom quand une ligne Suivi Tirage n'a pas de lot reconnu.
def fuzzy_key(value) -> str:
    text = (str(value) if value is not None else "").replace("\xa0", " ").upper()
    text = text.replace("-", " ")
    text = text.replace("PRERFECT", "PERFECT")
    # La fiche article colle souvent le nombre et l'unite ("200ML",
    # "140GRS") alors que l'Excel Suivi Tirage met un espace ("200 ML") -
    # on retire cet espace des deux cotes pour comparer sur la meme forme.
    text = re.sub(r"(\d)\s+(ML|GRS|G|L|KG)\b", r"\1\2", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not re.match(r"^-?\d+([.,]\d+)?$", text):
        return None
    try:
        return float(text.replace(",", "."))
    except ValueError:
        return None


# Les colonnes ARRET_* sont formatees comme des durees dans Excel (ex:
# "0:05"), pas comme de simples nombres - openpyxl les renvoie alors en
# datetime.time ou datetime.timedelta, jamais en float. On les convertit en
# fraction de journee (meme convention que les cellules deja numeriques,
# recuperee cote appli via excelFractionToMinutes = valeur * 24 * 60).
def to_day_fraction(value):
    if isinstance(value, timedelta):
        return value.total_seconds() / 86400
    if isinstance(value, time):
        seconds = value.hour * 3600 + value.minute * 60 + value.second
        return seconds / 86400
    if isinstance(value, datetime):
        seconds = value.hour * 3600 + value.minute * 60 + value.second
        return seconds / 86400
    return to_number(value)


def to_date_iso(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


# Les colonnes TEMPS OUVERT / FIN sont formatees comme heures dans Excel -
# openpyxl les renvoie en datetime.time (ou datetime.datetime si la cellule
# porte aussi une date). On les convertit en "HH:MM" pour matcher le format
# deja utilise par temps_demarage_lot/temps_arret_batch (input type="time").
def to_time_hhmm(value):
    if isinstance(value, time):
        return value.strftime("%H:%M")
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    if isinstance(value, (int, float)):
        total_minutes = round(float(value) * 24 * 60)
        hours, minutes = divmod(total_minutes % (24 * 60), 60)
        return f"{hours:02d}:{minutes:02d}"
    return None


def build_ligne_index(programme_lignes: list[dict]) -> dict[str, dict]:
    index: dict[str, dict] = {}
    for ligne in programme_lignes:
        numero_lot = ligne.get("numero_lot") or ""
        for code in numero_lot.split(","):
            key = normalize_code(code)
            if key:
                index[key] = ligne
    return index


def build_article_index(articles: list[dict]) -> dict[str, dict]:
    index: dict[str, dict] = {}
    for article in articles:
        key = fuzzy_key(article.get("nom_article"))
        if key:
            index[key] = article
    return index


def build_row_payload(
    row: tuple,
    row_number: int,
    source_fichier: str,
    ligne_by_code: dict[str, dict],
    article_by_id: dict[int, dict],
    article_by_fuzzy_name: dict[str, dict],
    stats: dict,
) -> dict | None:
    code_labo_raw = row[COL_CODE_LABO] if COL_CODE_LABO < len(row) else None
    code_labo = str(code_labo_raw).strip() if code_labo_raw not in (None, "") else ""

    if not code_labo:
        stats["sans_code_labo"] += 1
        return None

    matched_ligne = ligne_by_code.get(normalize_code(code_labo))

    excel_article_name = row[COL_ARTICLE] if COL_ARTICLE < len(row) else None

    zone = None
    chaine = None
    produit = None
    gamme = None
    type_article = None
    piece_par_carton = None
    qt_carton = to_number(row[COL_NBRE_CARTONS]) if COL_NBRE_CARTONS < len(row) else None
    programme_ligne_id = None

    if matched_ligne:
        stats["lies"] += 1
        programme_ligne_id = matched_ligne["id"]
        zone = matched_ligne.get("zone")
        chaine = matched_ligne.get("chaine")
        produit = matched_ligne.get("produit") or (
            str(excel_article_name).strip() if excel_article_name not in (None, "") else None
        )
        qt_carton = to_number(matched_ligne.get("qt_carton")) or qt_carton
        article_id = matched_ligne.get("article_id")
        article = article_by_id.get(article_id) if article_id else None
        if article:
            gamme = article.get("gamme")
            type_article = article.get("type_article")
            piece_par_carton = to_number(article.get("piece_par_carton"))
    else:
        stats["sans_lien"] += 1
        zone = str(row[COL_ZONE]).strip() if COL_ZONE < len(row) and row[COL_ZONE] not in (None, "") else None
        chaine = str(row[COL_LIGNE]).strip() if COL_LIGNE < len(row) and row[COL_LIGNE] not in (None, "") else None
        produit = (
            str(excel_article_name).strip() if excel_article_name not in (None, "") else None
        )
        article = article_by_fuzzy_name.get(fuzzy_key(excel_article_name))
        if article:
            gamme = article.get("gamme")
            type_article = article.get("type_article")
            piece_par_carton = to_number(article.get("piece_par_carton"))
        else:
            stats["article_non_trouve"] += 1

    programme_de_production_raw = row[COL_PROGRAMME_DE_PRODUCTION] if COL_PROGRAMME_DE_PRODUCTION < len(row) else None
    programme_de_production = (
        str(programme_de_production_raw).strip() if programme_de_production_raw not in (None, "") else None
    )

    return {
        "programme_ligne_id": programme_ligne_id,
        "code_labo": code_labo,
        "programme_de_production": programme_de_production,
        "date_jour": to_date_iso(row[COL_DATE]) if COL_DATE < len(row) else None,
        "zone": zone,
        "chaine": chaine,
        "produit": produit,
        "gamme": gamme,
        "type_article": type_article,
        "piece_par_carton": piece_par_carton,
        "chef_zone": str(row[COL_CHEF_ZONE]).strip() if COL_CHEF_ZONE < len(row) and row[COL_CHEF_ZONE] not in (None, "") else None,
        "chef_ligne": str(row[COL_CHEF_LIGNE]).strip() if COL_CHEF_LIGNE < len(row) and row[COL_CHEF_LIGNE] not in (None, "") else None,
        "tireur": str(row[COL_TIREUR]).strip() if COL_TIREUR < len(row) and row[COL_TIREUR] not in (None, "") else None,
        "cadence": to_number(row[COL_CADENCE_NET]) if COL_CADENCE_NET < len(row) else None,
        "poids_reel": to_number(row[COL_POIDS_MOYEN]) if COL_POIDS_MOYEN < len(row) else None,
        "qt_carton": qt_carton,
        "qt_fabriquer": to_number(row[COL_PRODUCTION]) if COL_PRODUCTION < len(row) else None,
        "arret_depot": to_day_fraction(row[COL_ARRET_DEPOT]) if COL_ARRET_DEPOT < len(row) else None,
        "arret_consommable_non_livre": to_day_fraction(row[COL_ARRET_CONSOMMABLE]) if COL_ARRET_CONSOMMABLE < len(row) else None,
        "arret_manque_conditionnement": to_day_fraction(row[COL_ARRET_MANQUE_CONDITIONNEMENT]) if COL_ARRET_MANQUE_CONDITIONNEMENT < len(row) else None,
        "arret_manque_vrac": to_day_fraction(row[COL_ARRET_MANQUE_VRAC]) if COL_ARRET_MANQUE_VRAC < len(row) else None,
        "arret_technique": to_day_fraction(row[COL_ARRET_TECHNIQUE]) if COL_ARRET_TECHNIQUE < len(row) else None,
        "arret_coupure_courant": to_day_fraction(row[COL_ARRET_COUPURE_COURANT]) if COL_ARRET_COUPURE_COURANT < len(row) else None,
        "arret_raclage_vrac": to_day_fraction(row[COL_ARRET_RACLAGE_VRAC]) if COL_ARRET_RACLAGE_VRAC < len(row) else None,
        "arret_changement_lot": to_day_fraction(row[COL_ARRET_CHANGEMENT_LOT]) if COL_ARRET_CHANGEMENT_LOT < len(row) else None,
        "arret_flacons_nc": to_day_fraction(row[COL_ARRET_FLACONS_NC]) if COL_ARRET_FLACONS_NC < len(row) else None,
        "arret_autre": to_day_fraction(row[COL_ARRET_AUTRE]) if COL_ARRET_AUTRE < len(row) else None,
        "temps_demarage_lot": to_time_hhmm(row[COL_TEMPS_OUVERT]) if COL_TEMPS_OUVERT < len(row) else None,
        "temps_arret_batch": to_time_hhmm(row[COL_FIN]) if COL_FIN < len(row) else None,
        "source_fichier": source_fichier,
    }


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    api_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not base_url or not api_key:
        print("Missing Supabase env values.")
        return 1

    if len(sys.argv) < 2:
        print("Usage: import_suivi_tirage.py <chemin_fichier.xlsx> [source_fichier]")
        return 1

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(f"Fichier introuvable: {workbook_path}")
        return 1

    source_fichier = sys.argv[2] if len(sys.argv) > 2 else workbook_path.name

    print("Lecture des programmes existants...")
    programme_lignes = fetch_all(
        base_url, api_key, "programme_lignes", "id,numero_lot,article_id,zone,chaine,produit,qt_carton"
    )
    ligne_by_code = build_ligne_index(programme_lignes)
    print(f"{len(programme_lignes)} programmes charges, {len(ligne_by_code)} codes de lot indexes.")

    print("Lecture des articles existants...")
    articles = fetch_all(base_url, api_key, "articles", "id,nom_article,gamme,type_article,piece_par_carton")
    article_by_id = {article["id"]: article for article in articles}
    article_by_fuzzy_name = build_article_index(articles)
    print(f"{len(articles)} articles charges.")

    workbook = load_workbook(workbook_path, read_only=True, data_only=True)

    # Nom de feuille recherche en tolerant espaces/casse differents d'un
    # fichier a l'autre (espace insecable, espace en fin de nom...).
    normalized_target = re.sub(r"\s+", " ", SHEET_NAME).strip().upper()
    matched_sheet_name = next(
        (
            name
            for name in workbook.sheetnames
            if re.sub(r"\s+", " ", name.replace("\xa0", " ")).strip().upper() == normalized_target
        ),
        None,
    )

    if not matched_sheet_name:
        print(f"Feuille '{SHEET_NAME}' introuvable. Feuilles disponibles : {workbook.sheetnames}")
        return 1

    print(f"Lecture de {workbook_path} ({matched_sheet_name})...")
    sheet = workbook[matched_sheet_name]

    stats = {"lies": 0, "sans_lien": 0, "sans_code_labo": 0, "article_non_trouve": 0}
    payloads: list[dict] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value in (None, "") for value in row):
            continue
        payload = build_row_payload(
            row, row_number, source_fichier, ligne_by_code, article_by_id, article_by_fuzzy_name, stats
        )
        if payload:
            payloads.append(payload)

    workbook.close()

    print(f"{len(payloads)} lignes a importer.")
    print(f"  - liees a un programme existant : {stats['lies']}")
    print(f"  - sans lien (importees quand meme) : {stats['sans_lien']}")
    print(f"    dont article non retrouve dans la fiche article : {stats['article_non_trouve']}")
    print(f"  - ignorees (pas de code labo) : {stats['sans_code_labo']}")

    chunk_size = 500
    for start in range(0, len(payloads), chunk_size):
        chunk = payloads[start : start + chunk_size]
        request_json(
            f"{base_url}/rest/v1/suivi_tirage",
            api_key,
            method="POST",
            payload=chunk,
            extra_headers={"Prefer": "return=minimal"},
        )
        print(f"  ... {min(start + chunk_size, len(payloads))}/{len(payloads)} envoyees")

    print("Import termine.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
