import json
import os
import re
import sys
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"P:\Partage-Depot-Produit-Fini\suivi DEPOT PF\proforma comande.xlsm")
IMPORT_COMMENT = "IMPORT_LISTE_COMMANDES_EXCEL"


def normalize_article(value: str) -> str:
    return (value or "").replace("\xa0", "").strip().upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line or line.startswith("NEXT_PUBLIC_SUPABASE_ANON_KEY"):
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def request_json(url: str, api_key: str, method: str = "GET", payload=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method=method,
    )

    with urllib.request.urlopen(request) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        return json.loads(raw) if raw else []


def request_no_content(url: str, api_key: str, method: str = "DELETE"):
    request = urllib.request.Request(
        url,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method=method,
    )

    with urllib.request.urlopen(request):
        return


def fetch_article_map(base_url: str, api_key: str) -> dict[str, int]:
    query = urllib.parse.urlencode(
        {
            "select": "id,article_normalise",
            "order": "id.asc",
            "limit": "5000",
        }
    )
    rows = request_json(f"{base_url}/rest/v1/articles?{query}", api_key)
    return {row["article_normalise"]: row["id"] for row in rows}


def fetch_previous_import_ids(base_url: str, api_key: str) -> list[int]:
    query = urllib.parse.urlencode(
        {
            "select": "id",
            "commentaire": f"eq.{IMPORT_COMMENT}",
            "limit": "5000",
        }
    )
    rows = request_json(f"{base_url}/rest/v1/commandes?{query}", api_key)
    return [int(row["id"]) for row in rows]


def clear_previous_import(base_url: str, api_key: str) -> None:
    commande_ids = fetch_previous_import_ids(base_url, api_key)
    if not commande_ids:
        return

    for commande_id in commande_ids:
        request_no_content(
            f"{base_url}/rest/v1/fifo_resultats?commande_id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )
        request_no_content(
            f"{base_url}/rest/v1/commande_lignes?commande_id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )
        request_no_content(
            f"{base_url}/rest/v1/commandes?id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )


def fetch_existing_command_ids_for_proformas(base_url: str, api_key: str, proformas: list[str]) -> list[int]:
    commande_ids: list[int] = []

    for proforma in proformas:
        query = urllib.parse.urlencode(
            {
                "select": "id",
                "numero_proforma": f"eq.{proforma}",
                "limit": "50",
            }
        )
        rows = request_json(f"{base_url}/rest/v1/commandes?{query}", api_key)
        commande_ids.extend(int(row["id"]) for row in rows)

    return commande_ids


def delete_commandes_by_ids(base_url: str, api_key: str, commande_ids: list[int]) -> None:
    for commande_id in sorted(set(commande_ids)):
        request_no_content(
            f"{base_url}/rest/v1/fifo_resultats?commande_id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )
        request_no_content(
            f"{base_url}/rest/v1/commande_lignes?commande_id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )
        request_no_content(
            f"{base_url}/rest/v1/commandes?id=eq.{commande_id}",
            api_key,
            method="DELETE",
        )


def sheet_sort_key(name: str):
    match = re.search(r"(\d+)", name)
    return int(match.group(1)) if match else 999999


def parse_nb_camion(value) -> int:
    try:
        nb_camion = int(value)
    except (TypeError, ValueError):
        return 1
    return nb_camion if nb_camion >= 1 else 1


def split_quantite_par_camion(quantite_totale: float, nb_camion: int) -> list[float]:
    """Split a total quantity (across all trucks) into nb_camion shares that
    add back up exactly to the total, distributing any remainder across the
    first few trucks instead of losing units to rounding."""
    base = quantite_totale // nb_camion
    remainder = quantite_totale - base * nb_camion
    return [base + (1 if truck_index < remainder else 0) for truck_index in range(nb_camion)]


def read_workbook_commands(workbook_path: Path, article_map: dict[str, int]):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = sorted(
        [name for name in workbook.sheetnames if name.strip().lower().startswith("comande")],
        key=sheet_sort_key,
    )

    commandes = []

    for sheet_name in sheet_names:
      ws = workbook[sheet_name]

      numero_proforma = ws["B1"].value
      client = ws["B2"].value
      nb_camion = parse_nb_camion(ws["D1"].value)
      genre_camion = ws["D2"].value

      if numero_proforma in (None, "") or client in (None, ""):
          continue

      lignes_totales = []

      for row_number in range(4, ws.max_row + 1):
          article_name = ws.cell(row_number, 1).value
          quantite = ws.cell(row_number, 2).value

          if article_name in (None, "") or quantite in (None, ""):
              continue

          article_key = normalize_article(str(article_name))
          article_id = article_map.get(article_key)
          if not article_id:
              continue

          lignes_totales.append({"article_id": article_id, "quantite_totale": float(quantite or 0)})

      if not lignes_totales:
          continue

      # One truck/container = one separate commande sharing the same base
      # numero_proforma (suffixed to satisfy the DB unique constraint), each
      # with its own share of the total quantity - mirrors how the web
      # "Ajouter commande" form creates one commande per truck, so an
      # imported order can also be delivered one truck at a time.
      base_proforma = str(numero_proforma).strip()
      per_truck_shares = {
          ligne["article_id"]: split_quantite_par_camion(ligne["quantite_totale"], nb_camion)
          for ligne in lignes_totales
      }

      for truck_index in range(nb_camion):
          truck_proforma = base_proforma if truck_index == 0 else f"{base_proforma}-{truck_index + 1}"

          truck_lignes = []
          for ligne in lignes_totales:
              qty_for_truck = per_truck_shares[ligne["article_id"]][truck_index]
              if qty_for_truck <= 0:
                  continue

              truck_lignes.append(
                  {
                      "article_id": ligne["article_id"],
                      "quantite_demandee": qty_for_truck,
                      "qt_non_dispo_total": 0,
                      "qt_non_dispo_2_mois": 0,
                      "qt_non_dispo_4_mois": 0,
                      "qt_non_dispo_6_mois": 0,
                  }
              )

          if not truck_lignes:
              continue

          commandes.append(
              {
                  "numero_proforma": truck_proforma,
                  "client": str(client).strip(),
                  "mode_chargement": str(genre_camion).strip() if genre_camion not in (None, "") else None,
                  "type_tc": str(nb_camion),
                  "statut": "EN_COURS",
                  "commentaire": IMPORT_COMMENT,
                  "lignes": truck_lignes,
              }
          )

    workbook.close()
    return commandes


def insert_commande(base_url: str, api_key: str, payload: dict) -> dict:
    rows = request_json(f"{base_url}/rest/v1/commandes", api_key, method="POST", payload=[payload])
    return rows[0]


def insert_commande_lignes(base_url: str, api_key: str, rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    return request_json(f"{base_url}/rest/v1/commande_lignes", api_key, method="POST", payload=rows)


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        print("Missing Supabase env values.")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK

    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    article_map = fetch_article_map(supabase_url, service_role_key)
    if not article_map:
        print("No articles found in database.")
        return 1

    commandes = read_workbook_commands(workbook_arg, article_map)
    if not commandes:
        print("No commands found in workbook.")
        return 1

    imported_proformas = [commande["numero_proforma"] for commande in commandes]
    imported_previous_ids = fetch_previous_import_ids(supabase_url, service_role_key)
    existing_same_proforma_ids = fetch_existing_command_ids_for_proformas(
        supabase_url, service_role_key, imported_proformas
    )
    delete_commandes_by_ids(
        supabase_url,
        service_role_key,
        imported_previous_ids + existing_same_proforma_ids,
    )

    imported_count = 0
    imported_lines = 0

    for commande in commandes:
        inserted_commande = insert_commande(
            supabase_url,
            service_role_key,
            {
                "numero_proforma": commande["numero_proforma"],
                "client": commande["client"],
                "mode_chargement": commande["mode_chargement"],
                "type_tc": commande["type_tc"],
                "statut": commande["statut"],
                "commentaire": commande["commentaire"],
            },
        )

        lignes_payload = [
            {
                "commande_id": inserted_commande["id"],
                **ligne,
            }
            for ligne in commande["lignes"]
        ]

        insert_commande_lignes(supabase_url, service_role_key, lignes_payload)
        imported_count += 1
        imported_lines += len(lignes_payload)

    print(f"Commandes importees: {imported_count}")
    print(f"Lignes importees: {imported_lines}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
