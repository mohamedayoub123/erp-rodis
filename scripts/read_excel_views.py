import json
import sys
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\ayoub\Desktop\MACRO EXCEL\GFPC-ENR-026 suivi stock  depot pf .xlsm")
FAMILY_SHEETS = [
    "White Secret",
    "Precious Perfect",
    "Perfect Glow",
    "BB Clear",
    "BB Clear VIT C",
    "Elixir",
    "Pro White",
    "Luxury Cocoa",
    "Luxury Avocado",
    "Egyptian Beauty",
    "MOROCCO SKIN",
    "ABSOLUTE CARE REALITY",
    "REAL CARE R",
    "TONE THERAPY R",
    "MY FAMILY CARE",
    "DERMATONE",
    "Coco Clear",
    "Cocoa Skin",
    "ECO+OFA+CDV+SKL",
    "SOOPURE",
    "EDT RODIS",
    "EDT REALITY",
    "MENTHOLE ETDIVERS",
]


def iso_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def workbook_path_from_argv() -> Path:
    if len(sys.argv) > 3:
        return Path(sys.argv[3])
    if len(sys.argv) > 2 and sys.argv[1] != "family_stock":
        return Path(sys.argv[2])
    return DEFAULT_WORKBOOK


def read_commandes_and_fifo(workbook_path: Path) -> dict:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws_cmd = wb["Comande"]
    ws_fifo = wb["Resulta"]

    commande = {
        "numero_proforma": str(ws_cmd["B1"].value).strip() if ws_cmd["B1"].value not in (None, "") else "",
        "client": str(ws_cmd["B2"].value).strip() if ws_cmd["B2"].value not in (None, "") else "",
        "lignes": [],
    }

    for row in range(4, ws_cmd.max_row + 1):
        article = ws_cmd.cell(row, 1).value
        if article in (None, ""):
            continue
        commande["lignes"].append(
            {
                "article": str(article).strip(),
                "quantite": float(ws_cmd.cell(row, 2).value or 0),
                "non_dispo_total": float(ws_cmd.cell(row, 3).value or 0),
                "non_dispo_2_mois": float(ws_cmd.cell(row, 4).value or 0),
                "non_dispo_4_mois": float(ws_cmd.cell(row, 5).value or 0),
                "non_dispo_6_mois": float(ws_cmd.cell(row, 6).value or 0),
            }
        )

    fifo_rows = []
    for row in range(4, ws_fifo.max_row + 1):
        article = ws_fifo.cell(row, 1).value
        code = ws_fifo.cell(row, 2).value
        if article in (None, "") or code in (None, ""):
            continue
        fifo_rows.append(
            {
                "article": str(article).strip(),
                "code": str(code).strip(),
                "quantite_chargee": float(ws_fifo.cell(row, 3).value or 0),
                "date": iso_date(ws_fifo.cell(row, 4).value),
                "chambre": str(ws_fifo.cell(row, 5).value).strip()
                if ws_fifo.cell(row, 5).value not in (None, "")
                else None,
            }
        )

    wb.close()
    return {"commande": commande, "fifo": fifo_rows}


def read_planning(workbook_path: Path) -> list[dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet_name in FAMILY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]

        for col in range(4, ws.max_column + 1):
            client = ws.cell(4, col).value
            nombre_camion = ws.cell(5, col).value
            mode_chargement = ws.cell(6, col).value
            numero_proforma = ws.cell(7, col).value

            if client in (None, "") and numero_proforma in (None, ""):
                continue

            for row in range(8, ws.max_row + 1):
                article = ws.cell(row, 1).value
                quantite = ws.cell(row, col).value
                if article in (None, "") or quantite in (None, "", 0):
                    continue

                rows.append(
                    {
                        "famille": sheet_name,
                        "client": str(client).strip() if client not in (None, "") else None,
                        "nombre_camion": float(nombre_camion or 0) if nombre_camion not in (None, "") else None,
                        "mode_chargement": str(mode_chargement).strip()
                        if mode_chargement not in (None, "")
                        else None,
                        "numero_proforma": str(numero_proforma).strip()
                        if numero_proforma not in (None, "")
                        else None,
                        "article": str(article).strip(),
                        "quantite_prevue": float(quantite or 0),
                    }
                )

    wb.close()
    return rows


def normalize_label(value) -> str:
    if value in (None, ""):
        return ""
    return str(value).replace("\xa0", " ").strip().upper()


def read_family_stock(workbook_path: Path, sheet_name: str) -> list[dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return []

    ws = wb[sheet_name]

    total_col = None
    stock_col = None
    reste_col = None

    for row in range(1, min(ws.max_row, 12) + 1):
        for col in range(1, ws.max_column + 1):
            label = normalize_label(ws.cell(row, col).value)
            if label == "TOTAL" and total_col is None:
                total_col = col
            elif label == "STOCK" and stock_col is None:
                stock_col = col
            elif label == "RESTE" and reste_col is None:
                reste_col = col

    rows: list[dict] = []

    for row in range(8, ws.max_row + 1):
        article = ws.cell(row, 1).value
        if article in (None, ""):
            continue

        article_name = str(article).replace("\xa0", " ").strip()
        if not article_name:
            continue

        rows.append(
            {
                "article": article_name,
                "total": float(ws.cell(row, total_col).value or 0) if total_col else 0,
                "stock": float(ws.cell(row, stock_col).value or 0) if stock_col else 0,
                "reste": float(ws.cell(row, reste_col).value or 0) if reste_col else 0,
            }
        )

    wb.close()
    return rows


def read_stock_dormant_sans_commande(workbook_path: Path) -> list[dict]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb["stock dorment sans comande"]
    rows: list[dict] = []

    for row in range(2, ws.max_row + 1):
        article = ws.cell(row, 1).value
        if article in (None, ""):
            continue
        rows.append(
            {
                "article": str(article).strip(),
                "type": str(ws.cell(row, 2).value).strip() if ws.cell(row, 2).value not in (None, "") else None,
                "gamme": str(ws.cell(row, 3).value).strip() if ws.cell(row, 3).value not in (None, "") else None,
                "marque": str(ws.cell(row, 4).value).strip() if ws.cell(row, 4).value not in (None, "") else None,
                "code": str(ws.cell(row, 5).value).strip() if ws.cell(row, 5).value not in (None, "") else None,
                "stock": float(ws.cell(row, 6).value or 0),
                "date": iso_date(ws.cell(row, 7).value),
            }
        )

    wb.close()
    return rows


def main() -> int:
    if len(sys.argv) < 2:
        print(json.dumps({"error": "Missing mode"}))
        return 1

    mode = sys.argv[1]
    workbook_path = workbook_path_from_argv()

    if not workbook_path.exists():
        print(json.dumps({"error": f"Workbook not found: {workbook_path}"}))
        return 1

    if mode == "commandes_fifo":
        result = read_commandes_and_fifo(workbook_path)
    elif mode == "planning":
        result = read_planning(workbook_path)
    elif mode == "family_stock":
        if len(sys.argv) < 3:
            print(json.dumps({"error": "Missing family sheet name"}))
            return 1
        result = read_family_stock(workbook_path, sys.argv[2])
    elif mode == "stock_dormant_sans_commande":
        result = read_stock_dormant_sans_commande(workbook_path)
    else:
        print(json.dumps({"error": f"Unknown mode: {mode}"}))
        return 1

    print(json.dumps(result, ensure_ascii=False))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
