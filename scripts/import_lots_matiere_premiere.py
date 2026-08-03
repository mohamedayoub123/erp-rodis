import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    r"P:\Partage-Moussa\GESTION DE DEPOT\Nouveau GesStock\gestion de stock - V1 5.xlsm"
)
DEFAULT_SHEET = "Etat_Lots"
CHUNK_SIZE = 100


def normalize_article(value: str) -> str:
    text = (value or "").replace("\xa0", "").strip()
    return text.upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def to_number(value) -> float | None:
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_date_str(value) -> str | None:
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return None


def clean_etat(value) -> str | None:
    if not value:
        return None
    text = str(value).strip()
    # La feuille source prefixe le statut d'un emoji (encode en "?" quand la
    # console ne le supporte pas) - on ne garde que le mot en clair.
    if "Perime" in text or "P\u00e9rim\u00e9" in text:
        return "Perime"
    if "Actif" in text:
        return "Actif"
    return text


def build_rows(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[DEFAULT_SHEET]

    rows_by_key: dict[str, dict] = {}

    # Colonnes : Designation, Lot, Date Reception, Date Expiration,
    # Mois en stock, Stock Actuel, Etat. Un seul lot par article dans cette
    # feuille (confirme : 676 lignes, 676 noms uniques).
    for row in sheet.iter_rows(min_row=2, values_only=True):
        designation = row[0]
        if not designation:
            continue

        article_normalise = normalize_article(str(designation))
        if not article_normalise:
            continue

        rows_by_key[article_normalise] = {
            "nom_article": str(designation).strip(),
            "article_normalise": article_normalise,
            "numero_lot": str(row[1]).strip() if row[1] else None,
            "date_reception": to_date_str(row[2]),
            "date_expiration": to_date_str(row[3]),
            "mois_en_stock": to_number(row[4]),
            "stock_actuel": to_number(row[5]),
            "etat": clean_etat(row[6]),
        }

    workbook.close()
    return list(rows_by_key.values())


def post_chunk(url: str, api_key: str, rows: list[dict]) -> None:
    query = urllib.parse.urlencode({"on_conflict": "article_normalise"})
    request = urllib.request.Request(
        f"{url}/rest/v1/lots_matiere_premiere?{query}",
        data=json.dumps(rows).encode("utf-8"),
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        method="POST",
    )

    with urllib.request.urlopen(request) as response:
        if response.status not in (200, 201, 204):
            raise RuntimeError(f"Unexpected response: {response.status}")


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    env_path = project_root / ".env.local"
    load_env(env_path)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get(
        "NEXT_PUBLIC_SUPABASE_ANON_KEY"
    )

    if not supabase_url:
        print("Missing NEXT_PUBLIC_SUPABASE_URL in .env.local")
        return 1

    if not service_role_key:
        print("Missing SUPABASE_SERVICE_ROLE_KEY in .env.local")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK

    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    rows = build_rows(workbook_arg)
    if not rows:
        print("No lot rows found in workbook.")
        return 1

    try:
        for start in range(0, len(rows), CHUNK_SIZE):
            post_chunk(supabase_url, service_role_key, rows[start : start + CHUNK_SIZE])
            print(f"Imported rows {start + 1} to {min(start + CHUNK_SIZE, len(rows))}")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        print(f"HTTP error {exc.code}: {body}")
        return 1
    except Exception as exc:
        print(f"Import failed: {exc}")
        return 1

    print(f"Import complete: {len(rows)} lot rows processed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
