import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


CHUNK_SIZE = 200


def normalize_client(value: str) -> str:
    return (value or "").replace("\xa0", "").strip().upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def build_rows(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active

    rows_by_key: dict[str, dict] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nom_client = row[0] if len(row) > 0 else None
        pays = row[1] if len(row) > 1 else None
        type_transport = row[2] if len(row) > 2 else None

        if not nom_client:
            continue

        client_normalise = normalize_client(str(nom_client))
        if not client_normalise:
            continue

        rows_by_key[client_normalise] = {
            "nom_client": str(nom_client).strip(),
            "client_normalise": client_normalise,
            "pays": str(pays).strip() if pays not in (None, "") else None,
            "mode_transport": str(type_transport).strip() if type_transport not in (None, "") else None,
        }

    workbook.close()
    return list(rows_by_key.values())


def post_chunk(url: str, api_key: str, rows: list[dict]) -> None:
    query = urllib.parse.urlencode({"on_conflict": "client_normalise"})
    request = urllib.request.Request(
        f"{url}/rest/v1/clients?{query}",
        data=json.dumps(rows).encode("utf-8"),
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        method="POST",
    )

    with urllib.request.urlopen(request) as response:
        if response.status not in (200, 201, 204):
            raise RuntimeError(f"Unexpected status {response.status}")


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        print("Missing Supabase env values.")
        return 1

    if len(sys.argv) < 2:
        print("Usage: import_clients_list.py <workbook_path>")
        return 1

    workbook_path = Path(sys.argv[1])
    if not workbook_path.exists():
        print(f"Workbook not found: {workbook_path}")
        return 1

    rows = build_rows(workbook_path)
    if not rows:
        print("No clients found in workbook.")
        return 1

    try:
        for start in range(0, len(rows), CHUNK_SIZE):
            post_chunk(supabase_url, service_role_key, rows[start : start + CHUNK_SIZE])
    except urllib.error.HTTPError as error:
        detail = error.read().decode("utf-8", errors="ignore")
        print(f"Erreur import clients: {error.code} {detail}")
        return 1

    print(f"Clients importes: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
