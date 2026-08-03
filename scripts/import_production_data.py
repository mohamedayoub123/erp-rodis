import json
import os
import re
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    r"P:\Partage-Production-Cosmetique\GFPC-ENR-032 Fabrication et Conditionement Cosmetique .xlsm"
)
DEFAULT_SHEET = "Data"
CHUNK_SIZE = 200


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def normalize_article(value: str) -> str:
    text = (value or "").replace("\xa0", "").strip()
    return text.upper()


# Cle de rapprochement plus tolerante que article_normalise : ignore les
# tirets/espaces multiples et corrige la faute de frappe connue
# "PRERFECT" (notre base) / "PERFECT" (cet Excel) pour eviter de creer des
# doublons - on garde toujours le nom EXISTANT de notre base, jamais celui
# de l'Excel, pour les articles deja presents.
def fuzzy_key(value: str) -> str:
    text = (value or "").replace("\xa0", " ").upper()
    text = text.replace("-", " ")
    text = text.replace("PRERFECT", "PERFECT")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def to_bool(value) -> bool:
    if value is None:
        return False
    text = str(value).strip().upper()
    return text.startswith("O")


def to_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, str) and not re.match(r"^-?\d+(\.\d+)?$", value.strip()):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def request_json(url: str, api_key: str, method: str = "GET", payload=None, extra_headers=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    headers = {
        "apikey": api_key,
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    if extra_headers:
        headers.update(extra_headers)

    request = urllib.request.Request(url, data=body, headers=headers, method=method)

    with urllib.request.urlopen(request) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        return json.loads(raw) if raw else []


def fetch_all_articles(base_url: str, api_key: str) -> list[dict]:
    rows: list[dict] = []
    from_idx = 0
    page_size = 1000

    while True:
        query = urllib.parse.urlencode({"select": "id,nom_article,article_normalise"})
        request = urllib.request.Request(
            f"{base_url}/rest/v1/articles?{query}",
            headers={
                "apikey": api_key,
                "Authorization": f"Bearer {api_key}",
                "Range": f"{from_idx}-{from_idx + page_size - 1}",
            },
        )
        with urllib.request.urlopen(request) as response:
            chunk = json.loads(response.read().decode("utf-8"))

        rows.extend(chunk)
        if len(chunk) < page_size:
            break
        from_idx += page_size

    return rows


COLUMN_MAP = [
    (8, "cadence", "number"),
    (9, "nb_carton_par_vrac", "number"),
    (10, "max_production_vrac_8h", "number"),
    (11, "contenance", "number"),
    (12, "nb_piece_par_max_vrac", "number"),
    (13, "piece_par_carton", "number"),
    (14, "min_vrac", "number"),
    (15, "max_vrac_auto", "number"),
    (16, "dispenseur_pcs_carton", "number"),
    (17, "besoin_pot_flacon", "bool"),
    (18, "besoin_capsule", "bool"),
    (19, "besoin_sleeve", "bool"),
    (20, "besoin_dispenseur", "bool"),
    (21, "besoin_carton", "bool"),
    (22, "besoin_etiquette", "bool"),
    (23, "besoin_etui", "bool"),
    (24, "code_auto", "text"),
    (25, "code_manu", "text"),
    (26, "vrac_max_manuel", "number"),
]


def build_production_fields(row: tuple) -> dict:
    fields: dict = {}
    for idx, field_name, kind in COLUMN_MAP:
        value = row[idx] if idx < len(row) else None
        if kind == "number":
            fields[field_name] = to_number(value)
        elif kind == "bool":
            fields[field_name] = to_bool(value)
        else:
            fields[field_name] = str(value).strip() if value not in (None, "") else None
    return fields


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    base_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    api_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not base_url or not api_key:
        print("Missing Supabase env values.")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    print("Lecture des articles existants...")
    existing = fetch_all_articles(base_url, api_key)
    by_fuzzy_key: dict[str, int] = {}
    for row in existing:
        by_fuzzy_key[fuzzy_key(row["nom_article"])] = row["id"]

    print(f"{len(existing)} articles existants charges.")

    workbook = load_workbook(workbook_arg, read_only=True, data_only=True, keep_vba=False)
    sheet = workbook[DEFAULT_SHEET]

    updates: list[dict] = []
    inserts: list[dict] = []
    inserted_names: list[str] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        nom_article = row[0]
        if not nom_article or not str(nom_article).strip():
            continue

        excel_name = str(nom_article).replace("\xa0", "").strip()
        key = fuzzy_key(excel_name)
        production_fields = build_production_fields(row)

        matched_id = by_fuzzy_key.get(key)

        if matched_id:
            updates.append({"id": matched_id, **production_fields})
        else:
            article_normalise = normalize_article(excel_name)
            inserts.append(
                {
                    "nom_article": excel_name,
                    "article_normalise": article_normalise,
                    "type_article": row[1] if row[1] else None,
                    "marque": row[2] if row[2] else None,
                    "gamme": row[3] if row[3] else None,
                    "min_stock": to_number(row[4]) or 0,
                    "max_stock": to_number(row[5]) or 0,
                    "volume_unitaire": to_number(row[6]),
                    "volume_stockage": to_number(row[7]),
                    **production_fields,
                }
            )
            inserted_names.append(excel_name)
            # Empeche un 2eme article "seulement dans l'Excel" avec le
            # meme nom de se dupliquer lui-meme dans cette meme execution.
            by_fuzzy_key[key] = -1

    workbook.close()

    print(f"{len(updates)} articles a mettre a jour (colonnes production).")
    print(f"{len(inserts)} nouveaux articles a ajouter.")

    # PATCH ligne par ligne : un upsert en masse (on_conflict=id) traite
    # chaque objet comme un INSERT potentiel et exige donc toutes les
    # colonnes NOT NULL (nom_article...), meme pour une ligne existante.
    for index, update in enumerate(updates, start=1):
        article_id = update["id"]
        fields = {k: v for k, v in update.items() if k != "id"}
        request_json(
            f"{base_url}/rest/v1/articles?id=eq.{article_id}",
            api_key,
            method="PATCH",
            payload=fields,
            extra_headers={"Prefer": "return=minimal"},
        )
        if index % 50 == 0 or index == len(updates):
            print(f"  ... {index}/{len(updates)} mis a jour")

    if inserts:
        request_json(
            f"{base_url}/rest/v1/articles?on_conflict=article_normalise",
            api_key,
            method="POST",
            payload=inserts,
            extra_headers={"Prefer": "resolution=merge-duplicates,return=minimal"},
        )

    print("\nNouveaux articles ajoutes:")
    for name in inserted_names:
        print(" -", name)

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
