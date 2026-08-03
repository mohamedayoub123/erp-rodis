import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\ayoub\Desktop\MACRO EXCEL\GFPC-ENR-026 suivi stock  depot pf .xlsm")
DEFAULT_SHEET = "entrer"
CHUNK_SIZE = 100


def normalize_article(value: str) -> str:
    return (value or "").replace("\xa0", "").strip().upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def iso_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def clean_meta(value):
    if value in (None, ""):
        return ""
    return str(value).replace("|", "/").strip()


def build_import_sortie_note(date_jour: str, client: str, bl: str, preparateur: str, quantite: float):
    if quantite <= 0 or (not client and not bl and not preparateur):
        return None

    return (
        f"SORTIEIMPORT|{date_jour}|-|{client or '-'}|{bl or '-'}|{preparateur or '-'}|{quantite}"
    )


def http_request(url: str, api_key: str, method: str = "GET", payload=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method=method,
    )

    with urllib.request.urlopen(request) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        return raw, response.status


def fetch_rows(base_url: str, api_key: str, table: str, params: dict[str, str]) -> list[dict]:
    query = urllib.parse.urlencode(params)
    raw, _ = http_request(f"{base_url}/rest/v1/{table}?{query}", api_key)
    return json.loads(raw or "[]")


def fetch_article_map(base_url: str, api_key: str) -> dict[str, int]:
    offset = 0
    article_map: dict[str, int] = {}

    while True:
        rows = fetch_rows(
            base_url,
            api_key,
            "articles",
            {
                "select": "id,article_normalise",
                "order": "id.asc",
                "limit": "1000",
                "offset": str(offset),
            },
        )
        if not rows:
            break

        for row in rows:
            article_map[row["article_normalise"]] = row["id"]

        offset += len(rows)

    return article_map


def fetch_imported_lot_ids(base_url: str, api_key: str) -> list[int]:
    lot_ids: list[int] = []
    offset = 0

    while True:
        rows = fetch_rows(
            base_url,
            api_key,
            "lots_stock",
            {
                "select": "id",
                "or": "(source_import.like.entrer:%,source_import.like.web:%)",
                "order": "id.asc",
                "limit": "1000",
                "offset": str(offset),
            },
        )
        if not rows:
            break

        lot_ids.extend(int(row["id"]) for row in rows if row.get("id") is not None)
        offset += len(rows)

    return lot_ids


def delete_fifo_rows_for_lots(base_url: str, api_key: str, lot_ids: list[int]) -> None:
    if not lot_ids:
        return

    for start in range(0, len(lot_ids), 200):
        chunk = lot_ids[start : start + 200]
        chunk_values = ",".join(str(lot_id) for lot_id in chunk)
        query = urllib.parse.urlencode({"lot_stock_id": f"in.({chunk_values})"})
        http_request(f"{base_url}/rest/v1/fifo_resultats?{query}", api_key, method="DELETE")


def clear_previous_import(base_url: str, api_key: str) -> None:
    imported_lot_ids = fetch_imported_lot_ids(base_url, api_key)
    delete_fifo_rows_for_lots(base_url, api_key, imported_lot_ids)

    query_import = urllib.parse.urlencode({"source_import": "like.entrer:%"})
    http_request(f"{base_url}/rest/v1/lots_stock?{query_import}", api_key, method="DELETE")

    query_web = urllib.parse.urlencode({"source_import": "like.web:%"})
    http_request(f"{base_url}/rest/v1/lots_stock?{query_web}", api_key, method="DELETE")


def build_rows(workbook_path: Path, article_map: dict[str, int]) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[DEFAULT_SHEET]

    rows: list[dict] = []

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        article_name = row[1]
        lot_number = row[4]

        if not article_name or lot_number in (None, ""):
            continue

        article_key = normalize_article(str(article_name))
        article_id = article_map.get(article_key)
        if not article_id:
            continue

        qte_entree = float(row[6] or 0)
        qte_sortie = float(row[7] or 0)

        if qte_entree == 0 and qte_sortie == 0:
            continue

        date_jour = iso_date(row[0]) or date.today().isoformat()
        date_fabrication = iso_date(row[5]) or date_jour
        client = clean_meta(row[10] if len(row) > 10 else None)
        bl = clean_meta(row[11] if len(row) > 11 else None)
        preparateur = clean_meta(row[14] if len(row) > 14 else None)
        note_lines: list[str] = []

        imported_sortie_note = build_import_sortie_note(
            date_jour,
            client,
            bl,
            preparateur,
            qte_sortie,
        )
        if imported_sortie_note:
            note_lines.append(imported_sortie_note)

        manual_note = str(row[15]).strip() if len(row) > 15 and row[15] not in (None, "") else None
        if manual_note:
            note_lines.append(manual_note)

        rows.append(
            {
                "article_id": article_id,
                "date_jour": date_jour,
                "numero_lot": str(lot_number).strip(),
                "code_normalise": str(lot_number).strip().upper(),
                "date_fabrication": date_fabrication,
                "qte_entree": qte_entree,
                "qte_sortie": qte_sortie,
                "chambre": str(row[13]).strip() if row[13] not in (None, "") else None,
                "code_pays": str(row[12]).strip() if row[12] not in (None, "") else None,
                "source_import": f"entrer:{row_number}",
                "note": "\n".join(note_lines) if note_lines else None,
            }
        )

    workbook.close()
    return rows


def post_chunk(base_url: str, api_key: str, rows: list[dict]) -> None:
    request = urllib.request.Request(
        f"{base_url}/rest/v1/lots_stock",
        data=json.dumps(rows).encode("utf-8"),
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )

    with urllib.request.urlopen(request) as response:
        if response.status not in (200, 201, 204):
            raise RuntimeError(f"Unexpected response: {response.status}")


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url:
        print("Missing NEXT_PUBLIC_SUPABASE_URL in .env.local")
        return 1

    if not service_role_key:
        print("Missing SUPABASE_SERVICE_ROLE_KEY in .env.local")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    article_map = fetch_article_map(supabase_url, service_role_key)
    if not article_map:
        print("No articles found in database. Import articles first.")
        return 1

    rows = build_rows(workbook_arg, article_map)
    if not rows:
        print("No stock lot rows found in workbook.")
        return 1

    try:
        clear_previous_import(supabase_url, service_role_key)
        print("Previous entrer and web stock cleared.")

        for start in range(0, len(rows), CHUNK_SIZE):
            post_chunk(supabase_url, service_role_key, rows[start : start + CHUNK_SIZE])
            print(f"Imported rows {start + 1} to {min(start + CHUNK_SIZE, len(rows))}")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        print(f"HTTP error {exc.code}: {body}")
        return 1
    except Exception as exc:
        print(f"Import failed: {exc}")
        return 1

    print(f"Import complete: {len(rows)} lot rows processed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
