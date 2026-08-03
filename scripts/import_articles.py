import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\ayoub\Desktop\MACRO EXCEL\GFPC-ENR-026 suivi stock  depot pf .xlsm")
DEFAULT_SHEET = "Data"
CHUNK_SIZE = 100


def normalize_article(value: str) -> str:
    text = (value or "").replace("\xa0", "").strip()
    return text.upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def build_rows(workbook_path: Path) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[DEFAULT_SHEET]

    rows_by_key: dict[str, dict] = {}

    for idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        nom_article = row[0]
        if not nom_article:
            continue

        article_normalise = normalize_article(str(nom_article))
        if not article_normalise:
            continue

        rows_by_key[article_normalise] = {
            "nom_article": str(nom_article).strip(),
            "article_normalise": article_normalise,
            "type_article": row[1] if row[1] else None,
            "marque": row[2] if row[2] else None,
            "gamme": row[3] if row[3] else None,
            "min_stock": float(row[4] or 0),
            "max_stock": float(row[5] or 0),
            "volume_unitaire": float(row[6]) if row[6] not in (None, "") else None,
            "volume_stockage": float(row[7]) if row[7] not in (None, "") else None,
        }

    workbook.close()
    return list(rows_by_key.values())


def post_chunk(url: str, api_key: str, rows: list[dict]) -> None:
    query = urllib.parse.urlencode({"on_conflict": "article_normalise"})
    request = urllib.request.Request(
        f"{url}/rest/v1/articles?{query}",
        data=json.dumps(rows).encode("utf-8"),
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "resolution=merge-duplicates,return=minimal",
        },
        method="POST",
    )

    with urllib.request.urlopen(request) as response:
        if response.status not in (200, 201, 204):
            raise RuntimeError(f"Unexpected response: {response.status}")


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    env_path = project_root / ".env.local"
    load_env(env_path)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get(
        "NEXT_PUBLIC_SUPABASE_ANON_KEY"
    )

    if not supabase_url:
        print("Missing NEXT_PUBLIC_SUPABASE_URL in .env.local")
        return 1

    if not service_role_key:
        print("Missing SUPABASE_SERVICE_ROLE_KEY in .env.local")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK

    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    rows = build_rows(workbook_arg)
    if not rows:
        print("No article rows found in workbook.")
        return 1

    try:
        for start in range(0, len(rows), CHUNK_SIZE):
            post_chunk(supabase_url, service_role_key, rows[start : start + CHUNK_SIZE])
            print(f"Imported rows {start + 1} to {min(start + CHUNK_SIZE, len(rows))}")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        print(f"HTTP error {exc.code}: {body}")
        return 1
    except Exception as exc:
        print(f"Import failed: {exc}")
        return 1

    print(f"Import complete: {len(rows)} article rows processed.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
