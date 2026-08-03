import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\ayoub\Desktop\MACRO EXCEL\GFPC-ENR-026 suivi stock  depot pf .xlsm")
DEFAULT_COMMANDE_SHEET = "Comande"
DEFAULT_FIFO_SHEET = "Resulta"


def normalize_article(value: str) -> str:
    return (value or "").replace("\xa0", "").strip().upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def iso_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return None


def request_json(url: str, api_key: str, method: str = "GET", payload=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method=method,
    )

    with urllib.request.urlopen(request) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        return json.loads(raw) if raw else []


def request_no_content(url: str, api_key: str, method: str = "DELETE"):
    request = urllib.request.Request(
        url,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method=method,
    )

    with urllib.request.urlopen(request):
        return


def fetch_article_map(base_url: str, api_key: str) -> dict[str, int]:
    query = urllib.parse.urlencode(
        {
            "select": "id,article_normalise",
            "order": "id.asc",
            "limit": "2000",
        }
    )
    rows = request_json(f"{base_url}/rest/v1/articles?{query}", api_key)
    return {row["article_normalise"]: row["id"] for row in rows}


def fetch_lot_map(base_url: str, api_key: str) -> dict[tuple[int, str], int]:
    offset = 0
    lot_map: dict[tuple[int, str], int] = {}

    while True:
        query = urllib.parse.urlencode(
            {
                "select": "id,article_id,code_normalise",
                "order": "id.asc",
                "limit": "1000",
                "offset": str(offset),
            }
        )
        rows = request_json(f"{base_url}/rest/v1/lots_stock?{query}", api_key)
        if not rows:
            break

        for row in rows:
            lot_map[(row["article_id"], row["code_normalise"])] = row["id"]

        offset += len(rows)

    return lot_map


def clear_previous_import(base_url: str, api_key: str) -> None:
    request_no_content(
        f"{base_url}/rest/v1/fifo_resultats?numero_lot=not.is.null",
        api_key,
        method="DELETE",
    )
    request_no_content(
        f"{base_url}/rest/v1/commandes?numero_proforma=not.is.null",
        api_key,
        method="DELETE",
    )


def insert_commande(base_url: str, api_key: str, payload: dict) -> dict:
    rows = request_json(f"{base_url}/rest/v1/commandes", api_key, method="POST", payload=[payload])
    return rows[0]


def insert_commande_lignes(base_url: str, api_key: str, rows: list[dict]) -> list[dict]:
    if not rows:
        return []
    return request_json(f"{base_url}/rest/v1/commande_lignes", api_key, method="POST", payload=rows)


def insert_fifo_rows(base_url: str, api_key: str, rows: list[dict]) -> None:
    if not rows:
        return
    request_json(f"{base_url}/rest/v1/fifo_resultats", api_key, method="POST", payload=rows)


def read_workbook_rows(workbook_path: Path, article_map: dict[str, int]):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    ws_cmd = workbook[DEFAULT_COMMANDE_SHEET]
    ws_fifo = workbook[DEFAULT_FIFO_SHEET]

    numero_proforma = str(ws_cmd["B1"].value).strip()
    client = str(ws_cmd["B2"].value).strip()

    commande_rows: list[dict] = []
    line_article_map: dict[int, int] = {}

    for row_number in range(4, ws_cmd.max_row + 1):
        article_name = ws_cmd.cell(row_number, 1).value
        if not article_name:
            continue

        article_key = normalize_article(str(article_name))
        article_id = article_map.get(article_key)
        if not article_id:
            continue

        quantite_demandee = float(ws_cmd.cell(row_number, 2).value or 0)
        commande_rows.append(
            {
                "article_id": article_id,
                "quantite_demandee": quantite_demandee,
                "qt_non_dispo_total": float(ws_cmd.cell(row_number, 3).value or 0),
                "qt_non_dispo_2_mois": float(ws_cmd.cell(row_number, 4).value or 0),
                "qt_non_dispo_4_mois": float(ws_cmd.cell(row_number, 5).value or 0),
                "qt_non_dispo_6_mois": float(ws_cmd.cell(row_number, 6).value or 0),
            }
        )
        line_article_map[row_number] = article_id

    fifo_rows: list[dict] = []
    for row_number in range(4, ws_fifo.max_row + 1):
        article_name = ws_fifo.cell(row_number, 1).value
        code = ws_fifo.cell(row_number, 2).value
        if not article_name or not code:
            continue

        article_key = normalize_article(str(article_name))
        article_id = article_map.get(article_key)
        if not article_id:
            continue

        fifo_rows.append(
            {
                "article_id": article_id,
                "numero_lot": str(code).strip(),
                "date_fabrication": iso_date(ws_fifo.cell(row_number, 4).value),
                "chambre": str(ws_fifo.cell(row_number, 5).value).strip()
                if ws_fifo.cell(row_number, 5).value not in (None, "")
                else None,
                "quantite_chargee": float(ws_fifo.cell(row_number, 3).value or 0),
            }
        )

    workbook.close()
    return numero_proforma, client, commande_rows, fifo_rows


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        print("Missing Supabase env values.")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    article_map = fetch_article_map(supabase_url, service_role_key)
    lot_map = fetch_lot_map(supabase_url, service_role_key)
    if not article_map:
        print("No articles in database. Import articles first.")
        return 1

    numero_proforma, client, commande_rows, fifo_rows = read_workbook_rows(workbook_arg, article_map)
    clear_previous_import(supabase_url, service_role_key)

    commande = insert_commande(
        supabase_url,
        service_role_key,
        {
            "numero_proforma": numero_proforma,
            "client": client,
            "date_commande": date.today().isoformat(),
            "statut": "IMPORTE",
            "commentaire": "Import automatique depuis feuille Comande",
        },
    )

    lignes_payload = []
    for row in commande_rows:
        lignes_payload.append({"commande_id": commande["id"], **row})

    inserted_lignes = insert_commande_lignes(supabase_url, service_role_key, lignes_payload)
    ligne_by_article: dict[int, int] = {}
    for line in inserted_lignes:
        ligne_by_article[line["article_id"]] = line["id"]

    fifo_payload = []
    ordre = 1
    for row in fifo_rows:
        lot_stock_id = lot_map.get((row["article_id"], row["numero_lot"].upper()))
        fifo_payload.append(
            {
                "commande_id": commande["id"],
                "commande_ligne_id": ligne_by_article.get(row["article_id"]),
                "article_id": row["article_id"],
                "lot_stock_id": lot_stock_id,
                "numero_lot": row["numero_lot"],
                "date_fabrication": row["date_fabrication"],
                "chambre": row["chambre"],
                "quantite_chargee": row["quantite_chargee"],
                "ordre_ligne": ordre,
                "regle_appliquee": "FIFO IMPORTE EXCEL",
            }
        )
        ordre += 1

    insert_fifo_rows(supabase_url, service_role_key, fifo_payload)
    print(f"Commande importee: {commande['numero_proforma']}")
    print(f"Lignes commande: {len(inserted_lignes)}")
    print(f"Lignes FIFO: {len(fifo_payload)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
