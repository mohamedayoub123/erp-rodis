import json
import os
import sys
import urllib.error
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(
    r"P:\Partage-Moussa\GESTION DE DEPOT\Nouveau GesStock\gestion de stock - V1 5.xlsm"
)
DEFAULT_SHEET = "Mouvements"
CHUNK_SIZE = 500
SOURCE_IMPORT = "excel:historique-mp"


def normalize_article(value: str) -> str:
    text = (value or "").replace("\xa0", "").strip()
    return text.upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def to_number(value):
    if value in (None, ""):
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def to_date_str(value):
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return None


def to_text(value):
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def fetch_all_articles(supabase_url: str, api_key: str) -> dict:
    by_normalise = {}
    offset = 0
    page_size = 1000

    while True:
        query = urllib.parse.urlencode(
            {
                "select": "id,nom_article",
                "order": "id.asc",
                "limit": page_size,
                "offset": offset,
            }
        )
        request = urllib.request.Request(
            f"{supabase_url}/rest/v1/articles_matiere_premiere?{query}",
            headers={
                "apikey": api_key,
                "Authorization": f"Bearer {api_key}",
                "User-Agent": "erp-rodis-import/1.0",
            },
        )
        with urllib.request.urlopen(request) as response:
            chunk = json.loads(response.read().decode("utf-8"))

        for row in chunk:
            by_normalise[normalize_article(row["nom_article"])] = row["id"]

        if len(chunk) < page_size:
            break
        offset += page_size

    return by_normalise


def build_rows(workbook_path: Path, article_by_normalise: dict) -> tuple[list[dict], int, int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[DEFAULT_SHEET]

    rows: list[dict] = []
    skipped_no_qty = 0
    skipped_no_article = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        (
            mvt_date,
            mvt_type,
            designation,
            quantite,
            unite,
            n_lot,
            _stock_restant,
            date_fab,
            date_exp,
            fourn_client,
            _emplacement,
            doss_erp,
            doss_4d,
            observation,
            _categorie,
            _etat,
        ) = row[:16]

        if not designation:
            continue

        qty = to_number(quantite)
        if qty is None or qty <= 0:
            skipped_no_qty += 1
            continue

        article_id = article_by_normalise.get(normalize_article(str(designation)))
        if article_id is None:
            skipped_no_article += 1
            continue

        type_lower = str(mvt_type or "").strip().lower()
        is_entree = type_lower.startswith("entr")

        numero_lot = to_text(n_lot)
        date_str = to_date_str(mvt_date)

        rows.append(
            {
                "article_id": article_id,
                "numero_lot": numero_lot,
                "code_normalise": numero_lot.upper() if numero_lot else None,
                "date_reception": date_str if is_entree else None,
                "date_fabrication": to_date_str(date_fab),
                "date_expiration": to_date_str(date_exp),
                "date_jour": date_str,
                "qte_entree": qty if is_entree else 0,
                "qte_sortie": 0 if is_entree else qty,
                "unite": to_text(unite),
                "fournisseur": to_text(fourn_client) if is_entree else None,
                "client": None if is_entree else to_text(fourn_client),
                "n_doss_erp": to_text(doss_erp),
                "n_doss_4d": to_text(doss_4d),
                "note": to_text(observation),
                "source_import": SOURCE_IMPORT,
            }
        )

    workbook.close()
    return rows, skipped_no_qty, skipped_no_article


def post_chunk(url: str, api_key: str, rows: list[dict]) -> None:
    request = urllib.request.Request(
        f"{url}/rest/v1/lots_stock_matiere_premiere",
        data=json.dumps(rows).encode("utf-8"),
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method="POST",
    )

    with urllib.request.urlopen(request) as response:
        if response.status not in (200, 201, 204):
            raise RuntimeError(f"Unexpected response: {response.status}")


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    env_path = project_root / ".env.local"
    load_env(env_path)

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY") or os.environ.get(
        "NEXT_PUBLIC_SUPABASE_ANON_KEY"
    )

    if not supabase_url:
        print("Missing NEXT_PUBLIC_SUPABASE_URL in .env.local")
        return 1

    if not service_role_key:
        print("Missing SUPABASE_SERVICE_ROLE_KEY in .env.local")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK

    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    print("Chargement des articles matiere premiere...")
    article_by_normalise = fetch_all_articles(supabase_url, service_role_key)
    print(f"{len(article_by_normalise)} articles charges.")

    print("Lecture de la feuille Mouvements...")
    rows, skipped_no_qty, skipped_no_article = build_rows(workbook_arg, article_by_normalise)
    print(
        f"{len(rows)} lignes a importer "
        f"(ignorees: {skipped_no_qty} qte invalide, {skipped_no_article} article non trouve)."
    )

    if not rows:
        print("Rien a importer.")
        return 1

    try:
        for start in range(0, len(rows), CHUNK_SIZE):
            chunk = rows[start : start + CHUNK_SIZE]
            post_chunk(supabase_url, service_role_key, chunk)
            print(f"Importe {start + len(chunk)} / {len(rows)}")
    except urllib.error.HTTPError as exc:
        body = exc.read().decode("utf-8", errors="ignore")
        print(f"HTTP error {exc.code}: {body}")
        return 1
    except Exception as exc:
        print(f"Import failed: {exc}")
        return 1

    print(f"Import complete: {len(rows)} lignes de mouvement importees.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
