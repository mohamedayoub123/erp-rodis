import json
import os
import sys
import urllib.parse
import urllib.request
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


DEFAULT_WORKBOOK = Path(r"C:\Users\ayoub\Desktop\MACRO EXCEL\GFPC-ENR-026 suivi stock  depot pf .xlsm")
FAMILY_SHEETS = [
    "White Secret",
    "Precious Perfect",
    "Perfect Glow",
    "BB Clear",
    "BB Clear VIT C",
    "Elixir",
    "Pro White",
    "Luxury Cocoa",
    "Luxury Avocado",
    "Egyptian Beauty",
    "MOROCCO SKIN",
    "ABSOLUTE CARE REALITY",
    "REAL CARE R",
    "TONE THERAPY R",
    "MY FAMILY CARE",
    "DERMATONE",
    "Coco Clear",
    "Cocoa Skin",
    "ECO+OFA+CDV+SKL",
    "SOOPURE",
    "EDT RODIS",
    "EDT REALITY",
    "MENTHOLE ETDIVERS",
]


def normalize_article(value: str) -> str:
    return (value or "").replace("\xa0", "").strip().upper()


def load_env(path: Path) -> None:
    if not path.exists():
        return

    for line in path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        key, value = line.split("=", 1)
        os.environ[key.strip()] = value.strip().strip('"').strip("'")


def request_json(url: str, api_key: str, method: str = "GET", payload=None):
    body = None if payload is None else json.dumps(payload).encode("utf-8")
    request = urllib.request.Request(
        url,
        data=body,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=representation",
        },
        method=method,
    )

    with urllib.request.urlopen(request) as response:
        raw = response.read().decode("utf-8", errors="ignore")
        return json.loads(raw) if raw else []


def request_no_content(url: str, api_key: str, method: str = "DELETE"):
    request = urllib.request.Request(
        url,
        headers={
            "apikey": api_key,
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json",
            "Prefer": "return=minimal",
        },
        method=method,
    )

    with urllib.request.urlopen(request):
        return


def fetch_map(base_url: str, api_key: str, table: str, key_field: str) -> dict[str, int]:
    query = urllib.parse.urlencode(
        {
            "select": f"id,{key_field}",
            "order": "id.asc",
            "limit": "3000",
        }
    )
    rows = request_json(f"{base_url}/rest/v1/{table}?{query}", api_key)
    return {str(row[key_field]).strip(): row["id"] for row in rows}


def clear_previous_import(base_url: str, api_key: str) -> None:
    request_no_content(
        f"{base_url}/rest/v1/famille_besoins?date_saisie=lte.{date.today().isoformat()}",
        api_key,
        method="DELETE",
    )


def build_rows(workbook_path: Path, article_map: dict[str, int], famille_map: dict[str, int]) -> list[dict]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    rows: list[dict] = []

    for sheet_name in FAMILY_SHEETS:
        if sheet_name not in workbook.sheetnames or sheet_name not in famille_map:
            continue

        ws = workbook[sheet_name]
        famille_id = famille_map[sheet_name]

        for col in range(4, ws.max_column + 1):
            client = ws.cell(4, col).value
            nombre_camion = ws.cell(5, col).value
            mode_chargement = ws.cell(6, col).value
            numero_proforma = ws.cell(7, col).value

            if client in (None, "") and numero_proforma in (None, ""):
                continue

            for row_number in range(8, ws.max_row + 1):
                article_name = ws.cell(row_number, 1).value
                qte = ws.cell(row_number, col).value

                if not article_name or qte in (None, "", 0):
                    continue

                article_id = article_map.get(normalize_article(str(article_name)))
                if not article_id:
                    continue

                rows.append(
                    {
                        "famille_id": famille_id,
                        "article_id": article_id,
                        "client": str(client).strip() if client not in (None, "") else None,
                        "nombre_camion": float(nombre_camion or 0) if nombre_camion not in (None, "") else None,
                        "mode_chargement": str(mode_chargement).strip()
                        if mode_chargement not in (None, "")
                        else None,
                        "numero_proforma": str(numero_proforma).strip()
                        if numero_proforma not in (None, "")
                        else None,
                        "quantite_prevue": float(qte or 0),
                        "qte_a_deduire_ak": 0,
                        "qte_a_deduire_al": 0,
                        "date_saisie": date.today().isoformat(),
                    }
                )

    workbook.close()
    return rows


def post_rows(base_url: str, api_key: str, rows: list[dict]) -> None:
    chunk_size = 500
    for start in range(0, len(rows), chunk_size):
        request_json(
            f"{base_url}/rest/v1/famille_besoins",
            api_key,
            method="POST",
            payload=rows[start : start + chunk_size],
        )


def main() -> int:
    project_root = Path(__file__).resolve().parents[1]
    load_env(project_root / ".env.local")

    supabase_url = os.environ.get("NEXT_PUBLIC_SUPABASE_URL")
    service_role_key = os.environ.get("SUPABASE_SERVICE_ROLE_KEY")

    if not supabase_url or not service_role_key:
        print("Missing Supabase env values.")
        return 1

    workbook_arg = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not workbook_arg.exists():
        print(f"Workbook not found: {workbook_arg}")
        return 1

    article_map = fetch_map(supabase_url, service_role_key, "articles", "article_normalise")
    famille_map = fetch_map(supabase_url, service_role_key, "familles", "nom_famille")
    rows = build_rows(workbook_arg, article_map, famille_map)
    clear_previous_import(supabase_url, service_role_key)
    post_rows(supabase_url, service_role_key, rows)
    print(f"Famille besoins importes: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
